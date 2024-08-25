import os
import re
import pandas as pd
import sqlalchemy as sa
import openpyxl as xl
from openpyxl.styles import PatternFill
import removed_for_confidentiality as g
from removed_for_confidentiality import Report, ReportTable
from removed_for_confidentiality import send_file_with_bot, BOT_IDS
from removed_for_confidentiality import GREEN_TO_RED_PALETTE
from removed_for_confidentiality import SlotsFactHorizont

DEFAULT_RENAME_COLS = {
    'MO_': 'МО',
    'FIL_': 'Филиал',
    'SERVICE_TYPE': 'Вид доступного ресурса',
    'SERVICE_TYPE_GROUP': 'Группа',
    'SERVICE_TYPE_GROUP_': 'Группа',
    'SUM_SLOTS': 'Всего слотов',
    'SUM_SLOTS_BUSY': 'Занято слотов',
    'BUSY_PERC': 'Процент занятых слотов',
}
DEFAULT_AR_CLAUSE = sa.and_(
    sa.literal_column('MO_TYPE').in_(['поликлиника']),
    sa.or_(
        sa.literal_column('SERVICE_TYPE_GROUP').in_(['1 уровень']),
    ),
)
DEFAULT_DATE_COL = 'START_DATE'
DEFAULT_GROUP_BY_COLS = [
    sa.literal_column('MO'),
    sa.literal_column('FIL'),
    sa.literal_column('SERVICE_TYPE'),
    sa.literal_column('SERVICE_TYPE_GROUP'),
    sa.literal_column('START_DATE'),
]
DEFAULT_QUOTA_TYPE = 'REC'
DEFAULT_SHEET_NAME = 'new'


# класс наследуется от родительского Report для доступа к свойствам _task_environment, tables, и методу add_table,
# которые стандартизируют пути к выходным файлам и работу с датафреймами
class SlotsRecsSFH(Report):
    def __init__(self, report_datetime, name: str = 'slots_recs_sfh'):
        """
        Формирует отчёт по слотам на горизонт от даты отчёта

        После записи файла можно вызвать один из методов для закраски ячеек:
        paint_red, paint_three_colors, paint_scale

        Свойства, которые можно (и нужно) переопределять:
        ar_clause: sqlalchemy.sql.elements.BooleanClauseList - срезы по ресурсам.
        можно срезать по всем колонкам, которые есть в MV_AVAILABLE_RESOURCES_SUPP
        group_by_cols: list - группировка - любые колонки из SFH_VIEW, должен содержать
        START_DATE или аналогичную колонку с датами
        date_col: str - по умолчанию START_DATE. Если в group_by_cols передали не START_DATE,
        а аналог, тут нужно указать новое название колонки с датами
        quota_type: str - по какой квоте из SFH_VIEW считать слоты: REC, REC_, FORWARD и т д
        rename_cols: dict[str, str] = словарь для переименования. нужно задавать если в group_by_cols
        добавили нестандартные колонки
        sort_out_df: callable - если нужна сортировка переопределяем метод, за основу взять
        определенный в классе метод sort_out_df
        sheet_name: str - пригождается если пишем несколько листов в один файл
        path: str - если нужно переопределить куда запишется файл. по умолчанию запишется в AUTO_REPORTS
        в папку с названием отчёта
        """
        super().__init__(name, report_datetime, datetime_format='%Y%m%d')
        self.xlsx_name = f'{self._name}_{self._task_environment.dates_for_names}.xlsx'
        self.sfh_filename = int(
            (report_datetime - pd.DateOffset(days=1)).strftime('%Y%m%d')
        )
        # check_partitions проверяет, что данные, необходимые для отчёта, загружены полностью
        SlotsFactHorizont(
            g.Period(report_datetime - pd.DateOffset(days=1))
        ).check_partitions()
        self.path = os.path.join(self._task_environment.xlsx_folder, self.xlsx_name)
        self.red_percent = 0.8
        self.ar_clause = DEFAULT_AR_CLAUSE
        self.group_by_cols = DEFAULT_GROUP_BY_COLS
        self.date_col = DEFAULT_DATE_COL
        self.quota_type = DEFAULT_QUOTA_TYPE
        self.rename_cols = DEFAULT_RENAME_COLS
        self.sheet_name = DEFAULT_SHEET_NAME

    def run(self):
        self.get_data()
        self.write()

    def get_data(self):
        """ Загрузить данные, обработать, записать необходимые датафреймы в свойство tables"""
        self.get_sfh_data()
        self.get_main_pivot()
        self.format_pivot_cols()
        self.get_piv_slash()
        self.get_df_total_cols()
        self.get_out_df()

    def write(self):
        """ Записать xlsx с отчётом """
        df_out_ = self.tables['df_out_'].df
        writer = pd.ExcelWriter(self.path, 'openpyxl')
        busy_perc_col = self.rename_cols['BUSY_PERC']
        writer = g.write_df(
            df_out_,
            writer=writer,
            red_more_format_col_nums=[
                df_out_.reset_index().columns.get_loc(busy_perc_col)
            ],
            sheet_name=self.sheet_name,
        )

    def paint_red(self):
        """Закрасить красным ячейки, где процент свободных слотов больше self.red_percent"""
        for cell, perc in self._get_slash_cells_values():
            if perc >= self.red_percent:
                cell.fill = PatternFill(fgColor='F8696B', fill_type='solid')
        self._wb.save(self.path)

    def paint_three_colors(self):
        """Закрасить ячейки в зависимости от процентра занятых слотов
        тремя цветами: желтым, ораньжевым, красным"""
        for cell, perc in self._get_slash_cells_values():
            if perc == 1:
                cell.fill = PatternFill(fgColor='F8696B', fill_type='solid')
            elif perc >= 0.95:
                cell.fill = PatternFill(fgColor='fba276', fill_type='solid')
            elif perc >= 0.85:
                cell.fill = PatternFill(fgColor='ede683', fill_type='solid')
        self._wb.save(self.path)

    def paint_scale(self):
        """Закрасить ячейки градиентом от зеленого до красного
        в зависимости от процентра занятых слотов"""
        for cell, perc in self._get_slash_cells_values():
            color = self.get_color(perc)
            cell.fill = PatternFill(fgColor=color, fill_type='solid')
        self._wb.save(self.path)

    def get_sfh_data(self):
        sfh_select = (
            sa.select(
                [
                    *self.group_by_cols,
                    sa.literal_column('count(IS_BUSY)').label('CNT_SLOTS'),
                    sa.literal_column('sum(IS_BUSY)').label('CNT_SLOTS_BUSY'),
                ]
            )
            .select_from(sa.text('MVIEWS.SFH_VIEW'))
            .where(
                sa.and_(
                    sa.literal_column('IS_FACT') == 0,
                    sa.literal_column('IS_HORIZONT') == 1,
                    sa.literal_column(self.quota_type) == 1,
                    sa.literal_column('FILENAME') == self.sfh_filename,
                    self.ar_clause,
                ),
            )
            .group_by(*self.group_by_cols)
        )

        df_sfh = pd.read_sql(sfh_select, g.engine)
        self.add_table(
            ReportTable(df_sfh, 'df_sfh', 'исходные данные из бд slots_fact_horizont')
        )

    def get_main_pivot(self):
        df_sfh = self.tables['df_sfh'].df
        index_main = [
            col for col in df_sfh.columns if col not in ['CNT_SLOTS', 'CNT_SLOTS_BUSY']
        ]
        piv_main = (
            df_sfh.set_index(index_main)
            .unstack(self.date_col)
            .sort_index(axis=1)
            .fillna(0)
            .astype(int)
        )
        piv_main = g.get_df_with_sum_row(piv_main)
        self.add_table(ReportTable(piv_main, 'piv_main', 'общий свод по датам'))

    def format_pivot_cols(self):
        piv_main = self.tables['piv_main'].df
        new_index = [
            piv_main.columns.get_level_values(n)
            for n in range(piv_main.columns.nlevels - 1)
            if piv_main.columns.get_level_values(n).name != self.date_col
        ] + [
            piv_main.columns.get_level_values(self.date_col).map(
                lambda x: x.strftime('%d.%m')
            )
        ]
        new_index
        piv_main.columns = pd.MultiIndex.from_arrays(new_index)
        self.tables['piv_main'].df = piv_main

    def get_piv_slash(self):
        piv_main = self.tables['piv_main'].df
        piv_slash = (
            piv_main.loc[:, pd.IndexSlice['CNT_SLOTS', :]].astype(str)
            + "/"
            + piv_main.loc[:, pd.IndexSlice['CNT_SLOTS_BUSY', :]].astype(str).values
        ).droplevel(0, axis=1)
        self.add_table(ReportTable(piv_slash, 'piv_slash', 'свод по датам через слеш'))

    def get_df_total_cols(self):
        piv_main = self.tables['piv_main'].df
        df_sum_slots_busy = (
            piv_main.loc[:, pd.IndexSlice['CNT_SLOTS_BUSY', :]]
            .sum(axis=1)
            .to_frame()
            .rename(columns={0: 'SUM_SLOTS_BUSY'})
        )
        df_sum_slots = (
            piv_main.loc[:, pd.IndexSlice['CNT_SLOTS', :]]
            .sum(axis=1)
            .to_frame()
            .rename(columns={0: 'SUM_SLOTS'})
        )
        df_total_cols = df_sum_slots.join(df_sum_slots_busy, how='outer')
        df_total_cols['BUSY_PERC'] = (
            df_total_cols['SUM_SLOTS_BUSY'] / df_total_cols['SUM_SLOTS']
        )
        self.add_table(
            ReportTable(
                df_total_cols,
                'df_total_cols',
                'колонки с суммой слотов и общим процентом',
            )
        )

    def _sort_out_df(self):
        df_out = self.tables['df_out'].df
        df_out = pd.concat(
            [
                df_out.loc[df_out.index.get_level_values(-1) == 'Всего:'],
                df_out.loc[df_out.index.get_level_values(-1) != 'Всего:'].sort_values(
                    'BUSY_PERC', ascending=False
                ),
            ]
        )
        self.tables['df_out'].df = df_out

    def get_out_df(self):
        df_total_cols = self.tables['df_total_cols'].df
        piv_slash = self.tables['piv_slash'].df
        df_out = df_total_cols.join(piv_slash, how='outer')
        self.add_table(ReportTable(df_out, 'df_out', 'финальная дф'))
        self._sort_out_df()
        df_out_ = (
            self.tables['df_out']
            .df.rename(columns=self.rename_cols)
            .rename_axis(index=self.rename_cols)
        )
        self.add_table(ReportTable(df_out_, 'df_out_', 'финальная дф переименованная'))

    def _get_slash_cells_values(self):
        self._wb = xl.load_workbook(self.path)
        try:
            worksheet = self._wb.get_sheet_by_name(self.sheet_name)
        except KeyError as e:
            print(e)
            return
        slash_cell_pattern = re.compile('^\d+\/\d+$')
        for row in worksheet.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if not type(cell.value) == str:
                    continue
                if slash_cell_pattern.match(cell.value):
                    slots_, recs_ = [int(el) for el in cell.value.split('/')]
                    if slots_ == 0:
                        continue
                    perc_ = recs_ / slots_
                    yield cell, perc_

    def _get_paint_coordinates(self):
        piv_busy_perc = self.tables['piv_busy_perc'].df
        df_out = self.tables['df_out'].df
        coordinates = []
        for col_ in range(piv_busy_perc.shape[0]):
            for row_ in range(piv_busy_perc.shape[1]):
                value = piv_busy_perc.iat[col_, row_]
                if pd.notnull(value):
                    coordinates.append(
                        [
                            (
                                col_
                                + df_out.columns.nlevels,  # кортеж с координатами ячеек в экселе, поправк на индес и заголовок
                                row_
                                + df_out.index.nlevels
                                + 3,  # +3 для колонок с суммой слотов
                            ),
                            value,
                        ]
                    )
        return coordinates

    def send(self):
        send_file_with_bot(
            BOT_IDS['CHAT_ID'],
            file_to_send=self.path,
            text_to_send=f'#{self._name}\n{self.task_environment.dates_for_names}',
        )

    @staticmethod
    def get_color(percent):
        index = int((round(percent, 2) * 10))
        color = GREEN_TO_RED_PALETTE[index]
        return color


if __name__ == '__main__':
    sfh = SlotsRecsSFH(pd.Timestamp(2023, 8, 15))
